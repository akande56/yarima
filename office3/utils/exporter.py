# # office3/utils/exporter.py

# from openpyxl import Workbook
# from openpyxl.styles import Font, Alignment, PatternFill
# from openpyxl.utils import get_column_letter
# from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
# from decimal import Decimal
# from collections import defaultdict
# from django.utils import timezone
# from django.http import HttpResponse
# from core.models import MineralBatch
# import re
# from datetime import datetime


# class TransactionExcelExporter:
#     def __init__(self):
#         # Full headers — exactly as in your original
#         self.headers = [
#             'Batch No.',
#             'Mineral',
#             'Grade',
#             'Supplier Name',
#             'Supplier Phone',
#             'Weight',
#             'Weight Unit',
#             'Agreed Price per Unit (₦)',
#             'Total Agreed Payout (₦)',
#             'Standard Price per Unit (₦)',
#             'Calculated Value (₦)',
#             'Total Value (₦)',
#             'Status',
#             'Date Received',
#             'Time Received (12Hr)',
#             'Recorded By',
#             'Payment Method',
#             'Payout Account Number',
#             'Payout Account Name',
#             'Payout Reference',
#             'Paid At Date',
#             'Paid At Time (12Hr)',
#             'Paid By',
#             'Approved By',
#             'Note',
#         ]

#         # Fixed column widths for data table (columns A to Y)
#         # These will be overridden for KPI columns A and B later
#         self.column_widths = [
#             14,  # [0]  Batch No
#             20,  # [1]  Mineral
#             15,  # [2]  Grade
#             22,  # [3]  Supplier Name
#             16,  # [4]  Supplier Phone
#             14,  # [5]  Weight
#             12,  # [6]  Weight Unit
#             24,  # [7]  Agreed Price per Unit
#             26,  # [8]  Total Agreed Payout
#             28,  # [9]  Standard Price per Unit
#             22,  # [10] Calculated Value
#             24,  # [11] Total Value (₦)
#             16,  # [12] Status
#             15,  # [13] Date Received
#             18,  # [14] Time Received
#             20,  # [15] Recorded By
#             24,  # [16] Payment Method
#             22,  # [17] Payout Account Number
#             22,  # [18] Payout Account Name
#             22,  # [19] Payout Reference
#             15,  # [20] Paid At Date
#             18,  # [21] Paid At Time
#             20,  # [22] Paid By
#             20,  # [23] Approved By
#             30,  # [24] Note
#         ]

#         # Styling
#         self.bold_font = Font(bold=True)
#         self.center_aligned_text = Alignment(horizontal="center")
#         self.title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

#     def export_to_response(self, queryset):
#         """
#         Export queryset to Excel with full fidelity, local time, safe decimals, and formatted numbers.
#         """
#         if not queryset.exists():
#             return HttpResponse(
#                 "No mineral batches match the filters.",
#                 content_type="text/plain",
#                 status=404
#             )

#         # Generate filename with date range
#         first = queryset.order_by('timestamp').first()
#         last = queryset.order_by('timestamp').last()
#         start_date = first.timestamp.date() if first else timezone.localdate()
#         end_date = last.timestamp.date() if last else timezone.localdate()
#         timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
#         filename = f'mineral_transactions_{start_date}_{end_date}_{timestamp}.xlsx'

#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         wb = self._create_workbook(queryset)
#         wb.save(response)
#         return response

#     def _create_workbook(self, queryset):
#         wb = Workbook()
#         wb.remove(wb.active)  # Remove default sheet

#         # === Sheet 1: All Transactions ===
#         ws_all = wb.create_sheet(title="All Transactions")
#         self._write_sheet_with_kpi(ws_all, queryset, is_main_sheet=True)

#         # === One sheet per mineral type ===
#         by_mineral = defaultdict(list)
#         for batch in queryset:
#             for item in batch.items.all():
#                 mineral_name = item.mineral_type.name if item.mineral_type else "Unknown Mineral"
#                 mineral_name = re.sub(r'[\\/?:*\[\]]', '_', mineral_name).strip()
#                 by_mineral[mineral_name].append((batch, item))

#         used_sheet_names = set()
#         for mineral_name, items in by_mineral.items():
#             safe_name = self._make_unique_sheet_name(mineral_name, used_sheet_names)
#             ws = wb.create_sheet(title=safe_name)
#             self._write_sheet_with_kpi(ws, queryset, items_filter=items, sheet_title=mineral_name)

#         return wb

#     def _write_sheet_with_kpi(self, ws, queryset, items_filter=None, sheet_title=None, is_main_sheet=False):
#         """Write KPI + headers + data + summary row with number formatting"""
#         # === KPI Summary ===
#         total_kg = Decimal('0.00')
#         total_lb = Decimal('0.00')
#         total_value = Decimal('0.00')

#         items_iter = items_filter or [(b, i) for b in queryset for i in b.items.all()]

#         for batch, item in items_iter:
#             if item.weight_unit == 'kg':
#                 total_kg += item.weight
#             elif item.weight_unit == 'lb':
#                 total_lb += item.weight
#             total_value += item.total_value.quantize(Decimal('0.00'))

#         total_converted_kg = total_kg + (total_lb * Decimal('0.453592'))
#         total_converted_lb = total_lb + (total_kg * Decimal('2.20462'))

#         ws.append(["KPI Summary"])
#         cell = ws.cell(row=ws.max_row, column=1)
#         cell.font = self.bold_font

#         ws.append(["Metric", "Value"])
#         header_row = ws.max_row
#         for col in [1, 2]:
#             cell = ws.cell(row=header_row, column=col)
#             cell.font = self.bold_font
#             cell.fill = self.title_fill
#             cell.alignment = self.center_aligned_text

#         ws.append(["Weight taken in kilograms (kg)", round(total_kg, 2)])
#         ws.append(["Weight taken in pounds (lb)", round(total_lb, 2)])
#         ws.append(["All weight converted to kilograms (kg)", round(total_converted_kg, 2)])
#         ws.append(["All weight converted to pounds (lb)", round(total_converted_lb, 2)])
#         ws.append(["Total monetary value (₦)", round(total_value, 2)])

#         ws.append([])  # Blank row before headers

#         # === Headers ===
#         ws.append(self.headers)

#         # Style header row
#         header_row = ws.max_row
#         for col_num, cell in enumerate(ws[header_row], 1):
#             cell.font = self.bold_font
#             cell.fill = self.title_fill
#             cell.alignment = self.center_aligned_text

#         # === Data Rows ===
#         for batch, item in items_iter:
#             row = self._format_row(batch, item)
#             ws.append(row)

#         # === Summary Row ===
#         summary_row = [""] * len(self.headers)
#         summary_row[1] = "TOTALS"
#         summary_row[5] = float(total_kg + total_lb)
#         summary_row[6] = "kg+lb"
#         summary_row[11] = float(total_value)

#         ws.append(summary_row)

#         # Style summary row
#         last_row = ws.max_row
#         for col_num in range(1, len(self.headers) + 1):
#             cell = ws.cell(row=last_row, column=col_num)
#             cell.font = self.bold_font
#             if col_num in [2, 6, 7, 11]:
#                 cell.fill = self.title_fill

#         # === Freeze Panes ===
#         ws.freeze_panes = 'A7'  # After KPI + blank + header

#         # === Apply Number Formatting (Commas for Money & Weight) ===
#         self._apply_number_formatting(ws)

#         # === Set Fixed Column Widths for Data Columns ===
#         for i, width in enumerate(self.column_widths, 1):
#             ws.column_dimensions[get_column_letter(i)].width = width

#         # === OVERRIDE: Widen KPI Columns A and B for Better Readability ===
#         ws.column_dimensions['A'].width = 42  # For long KPI labels
#         ws.column_dimensions['B'].width = 26  # For large formatted numbers

#         # Optional: Improve KPI label alignment
#         for row in range(3, 8):  # KPI metric rows (after header)
#             cell = ws.cell(row=row, column=1)
#             cell.alignment = Alignment(vertical="center")
#             cell = ws.cell(row=row, column=2)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#     def _format_row(self, batch, item):
#         """Format a single row with full precision and safe fallbacks"""
#         def format_time_12h(dt):
#             return timezone.localtime(dt).strftime('%I:%M:%S %p') if dt else ''

#         def format_date(dt):
#             return timezone.localtime(dt).strftime('%Y-%m-%d') if dt else ''

#         # Resolve mineral and grade
#         mineral_name = item.mineral_type.name if item.mineral_type else "Unknown Mineral"
#         grade_name = item.grade.grade_name if item.grade else "Unknown Grade"

#         # Standard pricing logic
#         if item.weight_unit == 'kg' and item.grade.price_per_kg is not None:
#             standard_unit_price = item.grade.price_per_kg
#             calculated_total = item.weight * standard_unit_price
#         elif item.weight_unit == 'lb' and item.grade.price_per_pound is not None:
#             standard_unit_price = item.grade.price_per_pound
#             calculated_total = item.weight * standard_unit_price
#         else:
#             standard_unit_price = Decimal('0.00')
#             calculated_total = Decimal('0.00')

#         # Agreed pricing
#         agreed_price_per_unit = item.agreed_payout
#         total_agreed_payout = (
#             agreed_price_per_unit * item.weight
#             if agreed_price_per_unit is not None else None
#         )

#         # Only show standard price if NOT negotiated
#         display_standard_unit_price = standard_unit_price if agreed_price_per_unit is None else None
#         display_calculated_value = calculated_total if agreed_price_per_unit is None else None

#         # Payment details
#         payment_methods = ", ".join(p.get_method_display() for p in batch.payments.all()[:2])
#         if batch.payments.count() > 2:
#             payment_methods += f" +{batch.payments.count() - 2} more"

#         payout_accounts = ", ".join(
#             p.payout_account_number for p in batch.payments.all() if p.payout_account_number
#         )
#         payout_names = ", ".join(
#             p.payout_account_name for p in batch.payments.all() if p.payout_account_name
#         )
#         references = ", ".join(
#             p.reference for p in batch.payments.all() if p.reference
#         )

#         # Clean text fields
#         supplier_name = ILLEGAL_CHARACTERS_RE.sub('?', batch.supplier_name or '')
#         notes = ILLEGAL_CHARACTERS_RE.sub(' ', batch.notes or '').replace('\n', ' ').replace('\r', ' ')

#         return [
#             batch.batch_no,
#             mineral_name,
#             grade_name,
#             supplier_name,
#             batch.supplier_phone or '',
#             float(item.weight) if item.weight else 0.0,
#             item.get_weight_unit_display(),
#             float(agreed_price_per_unit) if agreed_price_per_unit is not None else None,
#             float(total_agreed_payout) if total_agreed_payout is not None else None,
#             float(display_standard_unit_price) if display_standard_unit_price is not None else None,
#             float(display_calculated_value) if display_calculated_value is not None else None,
#             float(item.total_value) if item.total_value is not None else 0.0,
#             batch.get_status_display(),
#             format_date(batch.timestamp),
#             format_time_12h(batch.timestamp),
#             (batch.recorded_by.name or batch.recorded_by.username) if batch.recorded_by else '',
#             payment_methods or '',
#             payout_accounts or '',
#             payout_names or '',
#             references or '',
#             format_date(batch.paid_at),
#             format_time_12h(batch.paid_at),
#             (batch.paid_by.name or batch.paid_by.username) if batch.paid_by else '',
#             (batch.approved_by.name or batch.approved_by.username) if batch.approved_by else '',
#             notes,
#         ]

#     def _apply_number_formatting(self, ws):
#         """Apply number formatting to specific columns for readability"""
#         monetary_format = '#,##0.00'
#         weight_format = '#,##0.00'
#         date_format = 'YYYY-MM-DD'
#         time_format = 'hh:mm:ss AM/PM'

#         for row in ws.iter_rows(min_row=2):
#             # Weight (index 5)
#             cell = row[5]
#             if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
#                 cell.number_format = weight_format

#             # Monetary fields: indices 7 to 11
#             for col_idx in [7, 8, 9, 10, 11]:
#                 cell = row[col_idx]
#                 if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
#                     cell.number_format = monetary_format

#             # Date fields
#             for col_idx in [13, 20]:
#                 cell = row[col_idx]
#                 if cell.value and isinstance(cell.value, str) and '-' in cell.value:
#                     try:
#                         datetime.strptime(cell.value, '%Y-%m-%d')
#                         cell.number_format = date_format
#                     except ValueError:
#                         pass

#             # Time fields
#             for col_idx in [14, 21]:
#                 cell = row[col_idx]
#                 if cell.value and isinstance(cell.value, str) and ('AM' in cell.value.upper()):
#                     cell.number_format = time_format

#     def _make_unique_sheet_name(self, name, used_names):
#         """Make sheet name unique and Excel-safe (max 31 chars, no invalid chars)"""
#         invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
#         for char in invalid_chars:
#             name = name.replace(char, '_')
#         name = name.strip().strip('.')
#         if not name:
#             name = "Sheet"

#         candidate = name[:31]
#         counter = 1
#         original = candidate
#         while candidate in used_names:
#             suffix = f"_{counter}"
#             candidate = (original[:31 - len(suffix)] + suffix) if len(original) >= len(suffix) else suffix
#             counter += 1
#         used_names.add(candidate)
#         return candidate




# class SaleTransactionExcelExporter:
#     def __init__(self):
#         self.headers = [
#             'Sale ID',
#             'Reference No',
#             'Mineral',
#             'Grade',
#             'Buyer Name',
#             'Buyer Contact',
#             'Weight',
#             'Unit',
#             'Price per Unit (₦)',
#             'Total Value (₦)',
#             'Status',
#             'Sale Date',
#             'Sale Time (12Hr)',
#             'Processed By',
#             'Notes',
#             'BVN',
#             'NIN',
#             'International ID',
#             'Driver’s License',
#         ]

#         # Fixed column widths — optimized for clarity and performance
#         self.column_widths = [
#             12,  # Sale ID
#             16,  # Reference No
#             20,  # Mineral
#             15,  # Grade
#             20,  # Buyer Name
#             16,  # Buyer Contact
#             14,  # Weight
#             8,   # Unit
#             20,  # Price per Unit
#             20,  # Total Value
#             14,  # Status
#             14,  # Sale Date
#             16,  # Sale Time
#             18,  # Processed By
#             25,  # Notes
#             14,  # BVN
#             14,  # NIN
#             18,  # International ID
#             18,  # Driver’s License
#         ]

#         # Styling
#         self.bold_font = Font(bold=True)
#         self.center_aligned_text = Alignment(horizontal="center")
#         self.title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

#     def export_to_response(self, queryset):
#         """
#         Export filtered MineralSale queryset to Excel with date range in filename.
#         """
#         if not queryset.exists():
#             return HttpResponse(
#                 "No sales match the selected filters.",
#                 content_type="text/plain",
#                 status=404
#             )

#         # Determine date range for filename
#         first = queryset.order_by('sale_date').first()
#         last = queryset.order_by('sale_date').last()
#         start_date = first.sale_date.date() if first else timezone.localdate()
#         end_date = last.sale_date.date() if last else timezone.localdate()
#         timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
#         filename = f'sales_{start_date}_{end_date}_{timestamp}.xlsx'

#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         )
#         response['Content-Disposition'] = f'attachment; filename="{filename}"'

#         wb = self._create_workbook(queryset)
#         wb.save(response)
#         return response

#     def _create_workbook(self, queryset):
#         wb = Workbook()
#         wb.remove(wb.active)  # Remove default sheet

#         # Evaluate once
#         sales_list = list(queryset)

#         # === Sheet 1: All Sales ===
#         ws_all = wb.create_sheet(title="All Sales")
#         self._write_sheet_with_kpi(ws_all, sales_list, is_main_sheet=True)

#         # === Sheets per Mineral Type ===
#         by_mineral = defaultdict(list)
#         for sale in sales_list:
#             mineral_name = sale.mineral_type.name if sale.mineral_type else "Unknown Mineral"
#             mineral_name = re.sub(r'[\\/?:*\[\]]', '_', mineral_name).strip()
#             by_mineral[mineral_name].append(sale)

#         used_sheet_names = set()
#         for mineral_name, sales in by_mineral.items():
#             safe_name = self._make_unique_sheet_name(mineral_name, used_sheet_names)
#             ws = wb.create_sheet(title=safe_name)
#             self._write_sheet_with_kpi(ws, sales, sheet_title=mineral_name)

#         return wb

#     def _write_sheet_with_kpi(self, ws, sales_list, sheet_title=None, is_main_sheet=False):
#         """Write KPI block, headers, and data with formatting"""
#         # === KPI Calculation ===
#         total_kg = Decimal('0.00')
#         total_lb = Decimal('0.00')
#         total_value = Decimal('0.00')

#         for sale in sales_list:
#             if sale.quantity_unit == 'kg':
#                 total_kg += sale.quantity
#             elif sale.quantity_unit == 'lb':
#                 total_lb += sale.quantity
#             total_value += sale.total_price.quantize(Decimal('0.00'))

#         total_converted_kg = total_kg + (total_lb * Decimal('0.453592'))
#         total_converted_lb = total_lb + (total_kg * Decimal('2.20462'))

#         # === KPI Summary ===
#         ws.append(["KPI Summary"])
#         ws.cell(row=ws.max_row, column=1).font = self.bold_font

#         ws.append(["Metric", "Value"])
#         header_row = ws.max_row
#         for col in [1, 2]:
#             cell = ws.cell(row=header_row, column=col)
#             cell.font = self.bold_font
#             cell.fill = self.title_fill
#             cell.alignment = self.center_aligned_text

#         ws.append(["Weight sold in kilograms (kg)", round(total_kg, 2)])
#         ws.append(["Weight sold in pounds (lb)", round(total_lb, 2)])
#         ws.append(["All weight converted to kilograms (kg)", round(total_converted_kg, 2)])
#         ws.append(["All weight converted to pounds (lb)", round(total_converted_lb, 2)])
#         ws.append(["Total value (₦)", round(total_value, 2)])

#         ws.append([])  # Blank row
#         ws.append(self.headers)

#         # Style header row
#         header_row = ws.max_row
#         for cell in ws[header_row]:
#             cell.font = self.bold_font
#             cell.fill = self.title_fill
#             cell.alignment = self.center_aligned_text

#         # Add data rows
#         for sale in sales_list:
#             ws.append(self._format_row(sale))

#         # Freeze after KPI + blank + header
#         ws.freeze_panes = 'A7'

#         # Apply number formatting
#         self._apply_number_formatting(ws)

#         # Set fixed column widths
#         for i, width in enumerate(self.column_widths, 1):
#             ws.column_dimensions[get_column_letter(i)].width = width

#         # === OVERRIDE: Widen KPI Columns A and B for Readability ===
#         ws.column_dimensions['A'].width = 42  # Long labels
#         ws.column_dimensions['B'].width = 26  # Large numbers

#         # Align KPI values to the right
#         for row in range(3, 8):
#             cell = ws.cell(row=row, column=2)
#             cell.alignment = Alignment(horizontal="right", vertical="center")

#     def _format_row(self, sale):
#         """Format a single sale row with safe fallbacks and local time"""
#         def format_time_12h(dt):
#             return timezone.localtime(dt).strftime('%I:%M:%S %p') if dt else ''

#         def format_date(dt):
#             return timezone.localtime(dt).strftime('%Y-%m-%d') if dt else ''

#         mineral_name = sale.mineral_type.name if sale.mineral_type else "Unknown Mineral"
#         grade_name = sale.grade.grade_name if sale.grade else "Unknown Grade"
#         processed_by = (
#             sale.processed_by.name
#             or sale.processed_by.username
#             if sale.processed_by else ''
#         )

#         return [
#             sale.id,
#             sale.reference_number,
#             mineral_name,
#             grade_name,
#             sale.buyer_name,
#             sale.buyer_contact or '',
#             float(sale.quantity) if sale.quantity else 0.0,
#             sale.quantity_unit,
#             float(sale.price_per_unit) if sale.price_per_unit else 0.0,
#             float(sale.total_price) if sale.total_price else 0.0,
#             sale.get_status_display(),
#             format_date(sale.sale_date),
#             format_time_12h(sale.sale_date),
#             processed_by,
#             sale.notes or '',
#             sale.bvn or '',
#             sale.nin or '',
#             sale.international_id or '',
#             sale.driver_license or '',
#         ]

#     def _apply_number_formatting(self, ws):
#         """Apply number formatting: commas for money and weight"""
#         monetary_format = '#,##0.00'
#         weight_format = '#,##0.00'
#         date_format = 'YYYY-MM-DD'
#         time_format = 'hh:mm:ss AM/PM'

#         for row in ws.iter_rows(min_row=2):
#             # Weight (index 6)
#             cell = row[6]
#             if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
#                 cell.number_format = weight_format

#             # Price per Unit (index 8), Total Value (index 9)
#             for col_idx in [8, 9]:
#                 cell = row[col_idx]
#                 if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
#                     cell.number_format = monetary_format

#             # Sale Date (index 11)
#             cell = row[11]
#             if cell.value and isinstance(cell.value, str) and '-' in cell.value:
#                 try:
#                     timezone.datetime.strptime(cell.value, '%Y-%m-%d')
#                     cell.number_format = date_format
#                 except ValueError:
#                     pass

#             # Sale Time (index 12)
#             cell = row[12]
#             if cell.value and isinstance(cell.value, str) and ('AM' in cell.value.upper()):
#                 cell.number_format = time_format

#     def _make_unique_sheet_name(self, name, used_names):
#         """Ensure unique and valid Excel sheet name (max 31 chars)"""
#         invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
#         for char in invalid_chars:
#             name = name.replace(char, '_')
#         name = name.strip().strip('.')
#         if not name:
#             name = "Sheet"

#         candidate = name[:31]
#         counter = 1
#         original = candidate
#         while candidate in used_names:
#             suffix = f"_{counter}"
#             candidate = (original[:31 - len(suffix)] + suffix) if len(original) >= len(suffix) else suffix
#             counter += 1
#         used_names.add(candidate)
#         return candidate



# office3/utils/exporter.py

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from decimal import Decimal
from collections import defaultdict
from django.utils import timezone
from django.http import HttpResponse
from core.models import MineralBatch
import re
from datetime import datetime


class TransactionExcelExporter:
    def __init__(self):
        # Full headers — exactly as in your original
        self.headers = [
            'Batch No.',
            'Mineral',
            'Grade',
            'Supplier Name',
            'Supplier Phone',
            'Weight',
            'Weight Unit',
            'Agreed Price per Unit (₦)',
            'Total Agreed Payout (₦)',
            'Standard Price per Unit (₦)',
            'Calculated Value (₦)',
            'Total Value (₦)',
            'Status',
            'Date Received',
            'Time Received (12Hr)',
            'Recorded By',
            'Payment Method',
            'Payout Account Number',
            'Payout Account Name',
            'Payout Reference',
            'Paid At Date',
            'Paid At Time (12Hr)',
            'Paid By',
            'Approved By',
            'Note',
        ]

        # Fixed column widths for data table (columns A to Y)
        self.column_widths = [
            14,  # [0]  Batch No
            20,  # [1]  Mineral
            15,  # [2]  Grade
            22,  # [3]  Supplier Name
            16,  # [4]  Supplier Phone
            14,  # [5]  Weight
            12,  # [6]  Weight Unit
            24,  # [7]  Agreed Price per Unit
            26,  # [8]  Total Agreed Payout
            28,  # [9]  Standard Price per Unit
            22,  # [10] Calculated Value
            24,  # [11] Total Value (₦)
            16,  # [12] Status
            15,  # [13] Date Received
            18,  # [14] Time Received
            20,  # [15] Recorded By
            24,  # [16] Payment Method
            22,  # [17] Payout Account Number
            22,  # [18] Payout Account Name
            22,  # [19] Payout Reference
            15,  # [20] Paid At Date
            18,  # [21] Paid At Time
            20,  # [22] Paid By
            20,  # [23] Approved By
            30,  # [24] Note
        ]

        # Styling
        self.bold_font = Font(bold=True)
        self.center_aligned_text = Alignment(horizontal="center")
        self.title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    def export_to_response(self, queryset):
        """
        Export queryset to Excel with full fidelity, local time, safe decimals, and formatted numbers.
        """
        if not queryset.exists():
            return HttpResponse(
                "No mineral batches match the filters.",
                content_type="text/plain",
                status=404
            )

        # Generate filename with date range
        first = queryset.order_by('timestamp').first()
        last = queryset.order_by('timestamp').last()
        start_date = first.timestamp.date() if first else timezone.localdate()
        end_date = last.timestamp.date() if last else timezone.localdate()
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'mineral_transactions_{start_date}_{end_date}_{timestamp}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = self._create_workbook(queryset)
        wb.save(response)
        return response

    def _create_workbook(self, queryset):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # === Sheet 1: All Transactions ===
        ws_all = wb.create_sheet(title="All Transactions")
        self._write_sheet_with_kpi(ws_all, queryset, is_main_sheet=True)

        # === One sheet per mineral type ===
        by_mineral = defaultdict(list)
        for batch in queryset:
            for item in batch.items.all():
                mineral_name = item.mineral_type.name if item.mineral_type else "Unknown Mineral"
                mineral_name = re.sub(r'[\\/?:*\[\]]', '_', mineral_name).strip()
                by_mineral[mineral_name].append((batch, item))

        # ✅ Initialize used_sheet_names
        used_sheet_names = set()

        for mineral_name, items in by_mineral.items():
            safe_name = self._make_unique_sheet_name(mineral_name, used_sheet_names)
            ws = wb.create_sheet(title=safe_name)
            self._write_sheet_with_kpi(ws, queryset, items_filter=items, sheet_title=mineral_name)

        return wb

    def _write_sheet_with_kpi(self, ws, queryset, items_filter=None, sheet_title=None, is_main_sheet=False):
        """Write KPI + headers + data + summary row with number formatting"""
        # === KPI Summary ===
        total_kg = Decimal('0.00')
        total_lb = Decimal('0.00')
        total_value = Decimal('0.00')

        items_iter = items_filter or [(b, i) for b in queryset for i in b.items.all()]

        for batch, item in items_iter:
            if item.weight_unit == 'kg':
                total_kg += item.weight
            elif item.weight_unit == 'lb':
                total_lb += item.weight
            total_value += item.total_value.quantize(Decimal('0.00'))

        total_converted_kg = total_kg + (total_lb * Decimal('0.453592'))
        total_converted_lb = total_lb + (total_kg * Decimal('2.20462'))

        ws.append(["KPI Summary"])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = self.bold_font

        ws.append(["Metric", "Value"])
        header_row = ws.max_row
        for col in [1, 2]:
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.bold_font
            cell.fill = self.title_fill
            cell.alignment = self.center_aligned_text

        # Write KPI rows with number formatting
        kpi_data = [
            ("Weight taken in kilograms (kg)", round(total_kg, 2)),
            ("Weight taken in pounds (lb)", round(total_lb, 2)),
            ("All weight converted to kilograms (kg)", round(total_converted_kg, 2)),
            ("All weight converted to pounds (lb)", round(total_converted_lb, 2)),
            ("Total monetary value (₦)", round(total_value, 2)),
        ]

        for metric, value in kpi_data:
            ws.append([metric, value])
            cell = ws.cell(row=ws.max_row, column=2)  # Value column
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.append([])  # Blank row before headers

        # === Headers ===
        ws.append(self.headers)

        # Style header row
        header_row = ws.max_row
        for col_num, cell in enumerate(ws[header_row], 1):
            cell.font = self.bold_font
            cell.fill = self.title_fill
            cell.alignment = self.center_aligned_text

        # === Data Rows ===
        for batch, item in items_iter:
            row = self._format_row(batch, item)
            ws.append(row)

        # === Summary Row ===
        summary_row = [""] * len(self.headers)
        summary_row[1] = "TOTALS"
        summary_row[5] = float(total_kg + total_lb)
        summary_row[6] = "kg+lb"
        summary_row[11] = float(total_value)
        ws.append(summary_row)

        # Style summary row
        last_row = ws.max_row
        for col_num in range(1, len(self.headers) + 1):
            cell = ws.cell(row=last_row, column=col_num)
            cell.font = self.bold_font
            if col_num in [2, 6, 7, 11]:
                cell.fill = self.title_fill
            if col_num == 11:  # Total Value
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

        # === Freeze Panes ===
        ws.freeze_panes = 'A7'  # After KPI + blank + header

        # === Apply Number Formatting (Commas for Money & Weight) ===
        self._apply_number_formatting(ws)

        # === Set Fixed Column Widths for Data Columns ===
        for i, width in enumerate(self.column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # === OVERRIDE: Widen KPI Columns A and B for Better Readability ===
        ws.column_dimensions['A'].width = 42  # For long KPI labels
        ws.column_dimensions['B'].width = 26  # For large formatted numbers

        # Optional: Improve KPI label alignment
        for row in range(3, len(kpi_data) + 3):  # KPI value rows
            cell = ws.cell(row=row, column=1)
            cell.alignment = Alignment(vertical="center")
            cell = ws.cell(row=row, column=2)
            cell.alignment = Alignment(horizontal="right", vertical="center")

    def _format_row(self, batch, item):
        """Format a single row with full precision and safe fallbacks"""
        def format_time_12h(dt):
            return timezone.localtime(dt).strftime('%I:%M:%S %p') if dt else ''

        def format_date(dt):
            return timezone.localtime(dt).strftime('%Y-%m-%d') if dt else ''

        # Resolve mineral and grade
        mineral_name = item.mineral_type.name if item.mineral_type else "Unknown Mineral"
        grade_name = item.grade.grade_name if item.grade else "Unknown Grade"

        # Standard pricing logic
        if item.weight_unit == 'kg' and item.grade.price_per_kg is not None:
            standard_unit_price = item.grade.price_per_kg
            calculated_total = item.weight * standard_unit_price
        elif item.weight_unit == 'lb' and item.grade.price_per_pound is not None:
            standard_unit_price = item.grade.price_per_pound
            calculated_total = item.weight * standard_unit_price
        else:
            standard_unit_price = Decimal('0.00')
            calculated_total = Decimal('0.00')

        # Agreed pricing
        agreed_price_per_unit = item.agreed_payout
        total_agreed_payout = (
            agreed_price_per_unit * item.weight
            if agreed_price_per_unit is not None else None
        )

        # Only show standard price if NOT negotiated
        display_standard_unit_price = standard_unit_price if agreed_price_per_unit is None else None
        display_calculated_value = calculated_total if agreed_price_per_unit is None else None

        # Payment details
        payment_methods = ", ".join(p.get_method_display() for p in batch.payments.all()[:2])
        if batch.payments.count() > 2:
            payment_methods += f" +{batch.payments.count() - 2} more"

        payout_accounts = ", ".join(
            p.payout_account_number for p in batch.payments.all() if p.payout_account_number
        )
        payout_names = ", ".join(
            p.payout_account_name for p in batch.payments.all() if p.payout_account_name
        )
        references = ", ".join(
            p.reference for p in batch.payments.all() if p.reference
        )

        # Clean text fields
        supplier_name = ILLEGAL_CHARACTERS_RE.sub('?', batch.supplier_name or '')
        notes = ILLEGAL_CHARACTERS_RE.sub(' ', batch.notes or '').replace('\n', ' ').replace('\r', ' ')

        return [
            batch.batch_no,
            mineral_name,
            grade_name,
            supplier_name,
            batch.supplier_phone or '',
            float(item.weight) if item.weight else 0.0,
            item.get_weight_unit_display(),
            float(agreed_price_per_unit) if agreed_price_per_unit is not None else None,
            float(total_agreed_payout) if total_agreed_payout is not None else None,
            float(display_standard_unit_price) if display_standard_unit_price is not None else None,
            float(display_calculated_value) if display_calculated_value is not None else None,
            float(item.total_value) if item.total_value is not None else 0.0,
            batch.get_status_display(),
            format_date(batch.timestamp),
            format_time_12h(batch.timestamp),
            (batch.recorded_by.name or batch.recorded_by.username) if batch.recorded_by else '',
            payment_methods or '',
            payout_accounts or '',
            payout_names or '',
            references or '',
            format_date(batch.paid_at),
            format_time_12h(batch.paid_at),
            (batch.paid_by.name or batch.paid_by.username) if batch.paid_by else '',
            (batch.approved_by.name or batch.approved_by.username) if batch.approved_by else '',
            notes,
        ]

    def _apply_number_formatting(self, ws):
        """Apply number formatting to specific columns for readability"""
        monetary_format = '#,##0.00'
        weight_format = '#,##0.00'
        date_format = 'YYYY-MM-DD'
        time_format = 'hh:mm:ss AM/PM'

        for row in ws.iter_rows(min_row=2):
            # Weight (index 5)
            cell = row[5]
            if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
                cell.number_format = weight_format

            # Monetary fields: indices 7 to 11
            for col_idx in [7, 8, 9, 10, 11]:
                cell = row[col_idx]
                if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
                    cell.number_format = monetary_format

            # Date fields
            for col_idx in [13, 20]:
                cell = row[col_idx]
                if cell.value and isinstance(cell.value, str) and '-' in cell.value:
                    try:
                        datetime.strptime(cell.value, '%Y-%m-%d')
                        cell.number_format = date_format
                    except ValueError:
                        pass

            # Time fields
            for col_idx in [14, 21]:
                cell = row[col_idx]
                if cell.value and isinstance(cell.value, str) and ('AM' in cell.value.upper()):
                    cell.number_format = time_format

    def _make_unique_sheet_name(self, name, used_names):
        """Make sheet name unique and Excel-safe (max 31 chars, no invalid chars)"""
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '_')
        name = name.strip().strip('.')
        if not name:
            name = "Sheet"

        candidate = name[:31]
        counter = 1
        original = candidate
        while candidate in used_names:
            suffix = f"_{counter}"
            candidate = (original[:31 - len(suffix)] + suffix) if len(original) >= len(suffix) else suffix
            counter += 1
        used_names.add(candidate)
        return candidate


# ============ SALE EXPORTER ============

class SaleTransactionExcelExporter:
    def __init__(self):
        self.headers = [
            'Sale ID',
            'Reference No',
            'Mineral',
            'Grade',
            'Buyer Name',
            'Buyer Contact',
            'Weight',
            'Unit',
            'Price per Unit (₦)',
            'Total Value (₦)',
            'Status',
            'Sale Date',
            'Sale Time (12Hr)',
            'Processed By',
            'Notes',
            'BVN',
            'NIN',
            'International ID',
            'Driver’s License',
        ]

        # Fixed column widths — optimized for clarity and performance
        self.column_widths = [
            12,  # Sale ID
            16,  # Reference No
            20,  # Mineral
            15,  # Grade
            20,  # Buyer Name
            16,  # Buyer Contact
            14,  # Weight
            8,   # Unit
            20,  # Price per Unit
            20,  # Total Value
            14,  # Status
            14,  # Sale Date
            16,  # Sale Time
            18,  # Processed By
            25,  # Notes
            14,  # BVN
            14,  # NIN
            18,  # International ID
            18,  # Driver’s License
        ]

        # Styling
        self.bold_font = Font(bold=True)
        self.center_aligned_text = Alignment(horizontal="center")
        self.title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    def export_to_response(self, queryset):
        """
        Export filtered MineralSale queryset to Excel with date range in filename.
        """
        if not queryset.exists():
            return HttpResponse(
                "No sales match the selected filters.",
                content_type="text/plain",
                status=404
            )

        # Determine date range for filename
        first = queryset.order_by('sale_date').first()
        last = queryset.order_by('sale_date').last()
        start_date = first.sale_date.date() if first else timezone.localdate()
        end_date = last.sale_date.date() if last else timezone.localdate()
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'sales_{start_date}_{end_date}_{timestamp}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = self._create_workbook(queryset)
        wb.save(response)
        return response

    def _create_workbook(self, queryset):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Evaluate once
        sales_list = list(queryset)

        # === Sheet 1: All Sales ===
        ws_all = wb.create_sheet(title="All Sales")
        self._write_sheet_with_kpi(ws_all, sales_list, is_main_sheet=True)

        # === Sheets per Mineral Type ===
        by_mineral = defaultdict(list)
        for sale in sales_list:
            mineral_name = sale.mineral_type.name if sale.mineral_type else "Unknown Mineral"
            mineral_name = re.sub(r'[\\/?:*\[\]]', '_', mineral_name).strip()
            by_mineral[mineral_name].append(sale)

        # ✅ FIX: Initialize used_sheet_names
        used_sheet_names = set()

        for mineral_name, sales in by_mineral.items():
            safe_name = self._make_unique_sheet_name(mineral_name, used_sheet_names)
            ws = wb.create_sheet(title=safe_name)
            self._write_sheet_with_kpi(ws, sales, sheet_title=mineral_name)

        return wb

    def _write_sheet_with_kpi(self, ws, sales_list, sheet_title=None, is_main_sheet=False):
        """Write KPI block, headers, and data with formatting"""
        # === KPI Calculation ===
        total_kg = Decimal('0.00')
        total_lb = Decimal('0.00')
        total_value = Decimal('0.00')

        for sale in sales_list:
            if sale.quantity_unit == 'kg':
                total_kg += sale.quantity
            elif sale.quantity_unit == 'lb':
                total_lb += sale.quantity
            total_value += sale.total_price.quantize(Decimal('0.00'))

        total_converted_kg = total_kg + (total_lb * Decimal('0.453592'))
        total_converted_lb = total_lb + (total_kg * Decimal('2.20462'))

        # === KPI Summary ===
        ws.append(["KPI Summary"])
        ws.cell(row=ws.max_row, column=1).font = self.bold_font

        ws.append(["Metric", "Value"])
        header_row = ws.max_row
        for col in [1, 2]:
            cell = ws.cell(row=header_row, column=col)
            cell.font = self.bold_font
            cell.fill = self.title_fill
            cell.alignment = self.center_aligned_text

        # Write KPI rows with number formatting
        kpi_data = [
            ("Weight sold in kilograms (kg)", round(total_kg, 2)),
            ("Weight sold in pounds (lb)", round(total_lb, 2)),
            ("All weight converted to kilograms (kg)", round(total_converted_kg, 2)),
            ("All weight converted to pounds (lb)", round(total_converted_lb, 2)),
            ("Total value (₦)", round(total_value, 2)),
        ]

        for metric, value in kpi_data:
            ws.append([metric, value])
            cell = ws.cell(row=ws.max_row, column=2)  # Value column
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

        ws.append([])  # Blank row
        ws.append(self.headers)

        # Style header row
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = self.bold_font
            cell.fill = self.title_fill
            cell.alignment = self.center_aligned_text

        # Add data rows
        for sale in sales_list:
            ws.append(self._format_row(sale))

        # === Summary Row ===
        summary_row = [""] * len(self.headers)
        summary_row[1] = "TOTALS"
        summary_row[6] = float(total_kg + total_lb)
        summary_row[7] = "kg+lb"
        summary_row[9] = float(total_value)
        ws.append(summary_row)

        # Style summary row
        last_row = ws.max_row
        for col_num in range(1, len(self.headers) + 1):
            cell = ws.cell(row=last_row, column=col_num)
            cell.font = self.bold_font
            if col_num in [2, 6, 7, 9]:  # Mineral, Weight, Unit, Total Value
                cell.fill = self.title_fill
            if col_num == 9:  # Total Value
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

        # === Freeze Panes ===
        ws.freeze_panes = 'A7'

        # Apply number formatting
        self._apply_number_formatting(ws)

        # Set fixed column widths
        for i, width in enumerate(self.column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # === OVERRIDE: Widen KPI Columns A and B for Readability ===
        ws.column_dimensions['A'].width = 42  # Long labels
        ws.column_dimensions['B'].width = 26  # Large numbers

        # Align KPI values to the right
        for row in range(3, len(kpi_data) + 3):
            cell = ws.cell(row=row, column=2)
            cell.alignment = Alignment(horizontal="right", vertical="center")

    def _format_row(self, sale):
        """Format a single sale row with safe fallbacks and local time"""
        def format_time_12h(dt):
            return timezone.localtime(dt).strftime('%I:%M:%S %p') if dt else ''

        def format_date(dt):
            return timezone.localtime(dt).strftime('%Y-%m-%d') if dt else ''

        mineral_name = sale.mineral_type.name if sale.mineral_type else "Unknown Mineral"
        grade_name = sale.grade.grade_name if sale.grade else "Unknown Grade"
        processed_by = (
            sale.processed_by.name
            or sale.processed_by.username
            if sale.processed_by else ''
        )

        return [
            sale.id,
            sale.reference_number,
            mineral_name,
            grade_name,
            sale.buyer_name,
            sale.buyer_contact or '',
            float(sale.quantity) if sale.quantity else 0.0,
            sale.quantity_unit,
            float(sale.price_per_unit) if sale.price_per_unit else 0.0,
            float(sale.total_price) if sale.total_price else 0.0,
            sale.get_status_display(),
            format_date(sale.sale_date),
            format_time_12h(sale.sale_date),
            processed_by,
            sale.notes or '',
            sale.bvn or '',
            sale.nin or '',
            sale.international_id or '',
            sale.driver_license or '',
        ]

    def _apply_number_formatting(self, ws):
        """Apply number formatting: commas for money and weight"""
        monetary_format = '#,##0.00'
        weight_format = '#,##0.00'
        date_format = 'YYYY-MM-DD'
        time_format = 'hh:mm:ss AM/PM'

        for row in ws.iter_rows(min_row=2):
            # Weight (index 6)
            cell = row[6]
            if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
                cell.number_format = weight_format

            # Price per Unit (index 8), Total Value (index 9)
            for col_idx in [8, 9]:
                cell = row[col_idx]
                if isinstance(cell.value, (int, float, Decimal)) and cell.value not in (None, ''):
                    cell.number_format = monetary_format

            # Sale Date (index 11)
            cell = row[11]
            if cell.value and isinstance(cell.value, str) and '-' in cell.value:
                try:
                    timezone.datetime.strptime(cell.value, '%Y-%m-%d')
                    cell.number_format = date_format
                except ValueError:
                    pass

            # Sale Time (index 12)
            cell = row[12]
            if cell.value and isinstance(cell.value, str) and ('AM' in cell.value.upper()):
                cell.number_format = time_format

    def _make_unique_sheet_name(self, name, used_names):
        """Ensure unique and valid Excel sheet name (max 31 chars)"""
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid_chars:
            name = name.replace(char, '_')
        name = name.strip().strip('.')
        if not name:
            name = "Sheet"

        candidate = name[:31]
        counter = 1
        original = candidate
        while candidate in used_names:
            suffix = f"_{counter}"
            candidate = (original[:31 - len(suffix)] + suffix) if len(original) >= len(suffix) else suffix
            counter += 1
        used_names.add(candidate)
        return candidate
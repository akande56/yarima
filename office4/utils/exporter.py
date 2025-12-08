# office4/utils/exporter.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
from collections import defaultdict

class InvoiceExcelExporter:
    def __init__(self):
        self.headers = [
            'Invoice #', 'Expense Type', 'Vendor/Supplier', 'Payment Method',
            'Payment Date', 'Category', 'Amount (₦)', 'Description', 'Remark',
            'Created At', 'Updated At'
        ]
        self.bold_font = Font(bold=True)
        self.title_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    def export_to_response(self, queryset):
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'invoices_export_{timestamp}.xlsx'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb = self._create_workbook(queryset)
        wb.save(response)
        return response

    def _create_workbook(self, queryset):
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: All Invoices
        ws_all = wb.create_sheet(title="All Invoices")
        ws_all.append(self.headers)

        total_amount = Decimal('0.00')

        for inv in queryset:
            for item in inv.items.all():
                row = [
                    inv.sn,
                    inv.expense_type,
                    inv.vendor_supplier or '',
                    inv.get_payment_method_display() or '',
                    inv.payment_date,
                    item.get_category_display(),
                    float(item.amount),
                    inv.description or '',
                    inv.remark or '',
                    inv.created_at.strftime('%Y-%m-%d %H:%M'),
                    inv.updated_at.strftime('%Y-%m-%d %H:%M'),
                ]
                ws_all.append(row)
                total_amount += item.amount

        # Summary Row
        summary = ["", "TOTAL", "", "", "", "", float(total_amount)] + [""] * 4
        ws_all.append(summary)
        last_row = ws_all.max_row
        for col in range(1, 8):
            cell = ws_all.cell(row=last_row, column=col)
            cell.font = self.bold_font
            if col in [2, 7]:
                cell.fill = self.title_fill

        ws_all.freeze_panes = 'A2'
        self._auto_adjust_column_width(ws_all)

        # Group by Category
        by_category = defaultdict(list)
        for inv in queryset:
            for item in inv.items.all():
                by_category[item.get_category_display()].append((inv, item))

        for category, items in by_category.items():
            safe_name = self._make_sheet_name(category)
            ws = wb.create_sheet(title=safe_name)
            ws.append(["Category Summary"])
            ws.cell(row=ws.max_row, column=1).font = self.bold_font
            ws.append(["Metric", "Value"])
            for cell in ws[ws.max_row]:
                cell.font = self.bold_font
                cell.fill = self.title_fill

            cat_total = sum(item.amount for _, item in items)
            ws.append(["Total Expenses (₦)", float(cat_total)])
            ws.append([])

            ws.append(self.headers)
            for inv, item in items:
                row = [
                    inv.sn, inv.expense_type, inv.vendor_supplier or '',
                    inv.get_payment_method_display() or '', inv.payment_date,
                    item.get_category_display(), float(item.amount),
                    inv.description or '', inv.remark or '',
                    inv.created_at.strftime('%Y-%m-%d %H:%M'),
                    inv.updated_at.strftime('%Y-%m-%d %H:%M'),
                ]
                ws.append(row)

            kpi_rows = 4
            ws.freeze_panes = f'A{kpi_rows + 2}'
            self._auto_adjust_column_width(ws)

        return wb

    def _auto_adjust_column_width(self, ws):
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _make_sheet_name(self, name):
        invalid = ['\\', '/', '?', '*', '[', ']', ':']
        for char in invalid:
            name = name.replace(char, '_')
        return name.strip()[:31]